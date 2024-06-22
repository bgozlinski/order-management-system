import os
import h5py
from typing import Dict
from src.database.models import Order
from src.database.db import get_db
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import xml.etree.ElementTree as ET


def get_order_statistics() -> Dict[str, str]:
    """
    Retrieves statistics about the orders, such as the count of each status.

    Returns:
        Dict[str, str]: A dictionary with order status counts.
    """
    db = next(get_db())
    orders = db.query(Order).all()

    # Convert orders to a list of dictionaries and then to a DataFrame
    data = [order.to_dict() for order in orders]
    df = pd.DataFrame(data)

    # Count the occurrences of each status
    status_counts = df['status'].value_counts().to_dict()

    return status_counts


def generate_report_xlsx() -> str:
    """
    Generates an XLSX report containing all orders in the system.

    This function queries the database for all orders, converts the data to a DataFrame,
    creates an Excel workbook with a sheet containing the order data, and colors the rows
    based on the order status:
        - "New" orders are colored blue.
        - "In Progress" orders are colored yellow.
        - "Completed" orders are colored green.

    The generated report is saved in the 'reports' directory.

    Returns:
        str: The file path of the generated XLSX report.
    """
    db = next(get_db())
    orders = db.query(Order).all()

    # Convert orders to a list of dictionaries and then to a DataFrame
    data = [order.to_dict() for order in orders]
    if not data:
        raise ValueError("No orders found to generate report.")

    df = pd.DataFrame(data)

    # Create a new Excel workbook and add a sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Append headers to the sheet
    headers = df.columns.tolist()
    ws.append(headers)

    # Define colors for different order statuses
    status_colors = {
        "New": "0000FF",  # Blue
        "In Progress": "FFFF00",  # Yellow
        "Completed": "00FF00"  # Green
    }

    # Append data rows to the sheet and apply colors based on status
    for index, row in df.iterrows():
        ws.append(row.tolist())
        status = row['status']
        fill = PatternFill(start_color=status_colors.get(status, "FFFFFF"),
                           end_color=status_colors.get(status, "FFFFFF"), fill_type="solid")
        for cell in ws[index + 2]:
            cell.fill = fill

    # Define the reports directory and create it if it doesn't exist
    reports_dir = os.path.join(os.getcwd(), 'reports')
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)

    # Define the file path for the XLSX report
    report_path = os.path.join(reports_dir, "orders_report.xlsx")
    wb.save(report_path)

    return report_path


def export_orders_to_hdf5() -> str:
    """
    Export all orders to an HDF5 file.

    This function retrieves all orders from the database, converts the data into a DataFrame,
    converts datetime columns to strings, and saves the data to an HDF5 file in the 'reports' directory.

    Returns:
        str: The file path of the created HDF5 file.
    """
    db = next(get_db())
    orders = db.query(Order).all()

    data = [order.to_dict() for order in orders]
    df = pd.DataFrame(data)

    for column in df.select_dtypes(include=['datetime64[ns]']).columns:
        df[column] = df[column].astype(str)

    reports_dir = os.path.join(os.getcwd(), 'reports')
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)

    file_path = os.path.join(reports_dir, "orders.hdf5")

    with h5py.File(file_path, 'w') as f:
        for column in df.columns:
            f.create_dataset(column, data=df[column].values)

    return file_path


def import_orders_from_hdf5(file_path: str) -> None:
    """
    Import orders from an HDF5 file.

    This function reads order data from an HDF5 file, converts the data into a DataFrame,
    decodes byte strings, converts appropriate columns back to datetime, and merges the data into the database.

    Args:
        file_path (str): The file path of the HDF5 file to import.
    """
    db = next(get_db())

    with h5py.File(file_path, 'r') as f:
        data = {key: f[key][:] for key in f.keys()}

    df = pd.DataFrame(data)

    # Convert string columns back to datetime
    for column in df.columns:
        if df[column].dtype == object:
            try:
                df[column] = pd.to_datetime(df[column].astype(str), errors='ignore')
            except ValueError:
                pass

    for _, row in df.iterrows():
        order = Order(
            id=row['id'],
            name=row['name'],
            description=row['description'],
            creation_date=row['creation_date'],
            status=row['status']
        )
        db.merge(order)
    db.commit()


def export_orders_to_xml() -> str:
    """
    Export all orders to an XML file.

    This function retrieves all orders from the database and saves the data to an XML file in the 'reports' directory.

    Returns:
        str: The file path of the created XML file.
    """
    db = next(get_db())
    orders = db.query(Order).all()

    root = ET.Element("orders")
    for order in orders:
        order_elem = ET.Element("order")
        for key, value in order.to_dict().items():
            child = ET.Element(key)
            child.text = str(value)
            order_elem.append(child)
        root.append(order_elem)

    tree = ET.ElementTree(root)

    reports_dir = "reports"
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)

    file_path = os.path.join(reports_dir, "orders.xml")
    tree.write(file_path)

    return file_path


def import_orders_from_xml(file_path: str) -> None:
    """
    Import orders from an XML file.

    This function reads order data from an XML file and merges the data into the database.

    Args:
        file_path (str): The file path of the XML file to import.
    """
    db = next(get_db())

    tree = ET.parse(file_path)
    root = tree.getroot()

    for order_elem in root.findall('order'):
        order_data = {child.tag: child.text for child in order_elem}
        order_data['id'] = int(order_data['id'])
        order_data['creation_date'] = pd.to_datetime(order_data['creation_date'])

        order = Order(
            id=order_data['id'],
            name=order_data['name'],
            description=order_data['description'],
            creation_date=order_data['creation_date'],
            status=order_data['status']
        )
        db.merge(order)
    db.commit()
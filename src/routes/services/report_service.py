import os

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook

from src.database.db import get_db
from src.database.models import Order


def generate_report_xlsx() -> str:
    """
    Generates an XLSX report containing all orders in the system.

    This function queries the database for all orders, converts the data to a DataFrame,
    creates an Excel workbook with a sheet containing the order data, and colors the rows
    based on the order status:
        - "New" orders are colored blue.
        - "In Progress" orders are colored yellow.
        - "Completed" orders are colored green.

    The generated report is saved in the 'reports' directory.

    Returns:
        str: The file path of the generated XLSX report.
    """
    db = next(get_db())
    orders = db.query(Order).all()

    # Convert orders to a list of dictionaries and then to a DataFrame
    data = [order.to_dict() for order in orders]
    if not data:
        raise ValueError("No orders found to generate report.")

    df = pd.DataFrame(data)

    # Create a new Excel workbook and add a sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Append headers to the sheet
    headers = df.columns.tolist()
    ws.append(headers)

    # Define colors for different order statuses
    status_colors = {
        "New": "0000FF",  # Blue
        "In Progress": "FFFF00",  # Yellow
        "Completed": "00FF00"  # Green
    }

    # Append data rows to the sheet and apply colors based on status
    for index, row in df.iterrows():
        ws.append(row.tolist())
        status = row['status']
        fill = PatternFill(start_color=status_colors.get(status, "FFFFFF"),
                           end_color=status_colors.get(status, "FFFFFF"), fill_type="solid")
        for cell in ws[index + 2]:
            cell.fill = fill

    # Define the reports directory and create it if it doesn't exist
    reports_dir = os.path.join(os.getcwd(), 'reports')
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)

    # Define the file path for the XLSX report
    report_path = os.path.join(reports_dir, "orders_report.xlsx")
    wb.save(report_path)

    return report_path
